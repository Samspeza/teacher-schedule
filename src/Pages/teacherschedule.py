from multiprocessing import connection
from os import name
import sqlite3
import tkinter as tk
from tkinter import ttk
import random
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from tkinter import PhotoImage
from tkinter import filedialog
from tkinter import messagebox
from tkinter.messagebox import askyesno
import sys
import os

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../')))  
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'DbContext'))
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'UserControl'))
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'CSS'))
from DbContext.database import DB_NAME
from DbContext.models import get_teachers
from CSS.style import *
from UserControl.config import get_available_lab, get_class_course, get_class_id, get_discipline_id_by_name, get_disciplines, get_laboratories, get_teacher_availability_for_timetable, get_teacher_data, get_teacher_limits,  coordinator_id, teachers, teacher_limits, classes, days_of_week, time_slots
from ScreenManager import ScreenManager
from UserControl.sidebar import create_sidebar

class TimetableApp:
    def __init__(self, root, coordinator_id):
        self.manual_timetable = {}
        self.root = root
        self.coordinator_id = coordinator_id
        self.teachers = self.get_teachers(self.coordinator_id)
        self.root.title("Gerenciamento de Grade de Aulas")
        self.root.geometry("1175x900")
        self.root.config(bg=BACKGROUND_COLOR)
        self.selected_cell = None
        self.selected_grades = []

        self.teacher_allocations = {teacher: set() for teacher in self.teachers}

        self.main_frame = tk.Frame(self.root, bg=BACKGROUND_COLOR)
        self.main_frame.pack(fill="both", expand=True)

        self.setup_ui()

    def setup_ui(self):
        self.sidebar_frame = create_sidebar(
            self.main_frame, 
            show_home_screen=self.show_home_screen,
            show_modules_screen=self.show_modules_screen,
            save_changes=self.save_changes
        )
        self.action_frame = tk.Frame(self.main_frame, bg=BACKGROUND_COLOR)
        self.action_frame.pack(pady=10, padx=10, fill="x")

        # Cabeçalho com o título
        self.header_frame = tk.Frame(self.main_frame, bg="#F8F8F8")
        self.header_frame.pack(pady=20, fill="x", padx=10)

        title_label = tk.Label(
            self.header_frame,
            text="Criar Grade",
            font=("Arial", 16, "bold"),
            bg="#F8F8F8",
            fg="#2A72C3",
            cursor="hand2",
            anchor="w"
        )
        title_label.pack(fill="x", padx=10)

        # Frame para os botões de ação
        self.action_frame = tk.Frame(self.main_frame, bg=BACKGROUND_COLOR)
        self.action_frame.pack(pady=10, padx=10, fill="x")

        # Carregamento dos ícones dos botões
        self.save_icon = PhotoImage(file="icons/salvar.png").subsample(20, 20)
        self.cancel_icon = PhotoImage(file="icons/cancel.png").subsample(20, 20)
        self.list_icon = PhotoImage(file="icons/list.png").subsample(20, 20)
        self.edit_icon = PhotoImage(file="icons/edit.png").subsample(20, 20)
        self.delete_icon = PhotoImage(file="icons/delete.png").subsample(20, 20)
        self.create_icon = PhotoImage(file="icons/mais.png").subsample(20, 20)
        self.download_icon = PhotoImage(file="icons/download.png").subsample(20, 20)

        # Botões de ação dentro do frame action_frame
        self.create_button = tk.Button(
            self.action_frame,
            image=self.create_icon,
            command=self.create_manual_schedule,
            padx=8,
            pady=4,
            borderwidth=0,
            relief="flat",
            highlightthickness=0,
            activebackground=BACKGROUND_COLOR
        )
        self.create_button.pack(side="left", padx=8)

        self.list_button = tk.Button(
            self.action_frame,
            image=self.list_icon,
            command=self.show_timetable,
            padx=8,
            pady=4,
            borderwidth=0,
            relief="flat",
            highlightthickness=0,
            activebackground=BACKGROUND_COLOR
        )
        self.list_button.pack(side="left", padx=8)

        self.edit_button = tk.Button(
            self.action_frame,
            image=self.edit_icon,
            command=self.edit_cell,
            state="disabled",
            padx=8,
            pady=4,
            borderwidth=0,
            relief="flat",
            highlightthickness=0,
            activebackground=BACKGROUND_COLOR
        )
        self.edit_button.pack(side="left", padx=8)

        self.save_button = tk.Button(
            self.action_frame,
            image=self.save_icon,
            command=self.save_changes,
            state="disabled",
            padx=8,
            pady=4,
            borderwidth=0,
            relief="flat",
            highlightthickness=0,
            activebackground=BACKGROUND_COLOR
        )
        self.save_button.pack(side="left", padx=8)

        self.cancel_button = tk.Button(
            self.action_frame,
            image=self.cancel_icon,
            command=self.cancel_edit,
            state="disabled",
            padx=8,
            pady=4,
            borderwidth=0,
            relief="flat",
            highlightthickness=0,
            activebackground=BACKGROUND_COLOR
        )
        self.cancel_button.pack(side="left", padx=8)

        self.delete_button = tk.Button(
            self.action_frame,
            image=self.delete_icon,
            command=self.confirm_delete_schedule,
            state="disabled",
            padx=8,
            pady=4,
            borderwidth=0,
            relief="flat",
            highlightthickness=0,
        )
        self.delete_button.pack(side="left", padx=8)

        self.download_button = tk.Button(
            self.action_frame,
            image=self.download_icon,
            command=self.show_export_options, 
            state="disabled",
            padx=8,
            pady=4,
            borderwidth=0,
            relief="flat",
            highlightthickness=0,
        )
        self.download_button.pack(side="left", padx=8)

        self.lab_config_button = tk.Button(
            self.action_frame,
            text="Configurar Laboratórios",
            command=self.open_lab_config_window,
            padx=8,
            pady=4,
            bg="#DDEEFF",
            fg="#003366",
            font=("Arial", 10, "bold"),
            relief="raised"
        )
        self.lab_config_button.pack(side="left", padx=8)

        # Área para exibição da grade de aulas
        self.timetable_frame = tk.Frame(self.main_frame, bg=BACKGROUND_COLOR)
        self.timetable_frame.pack(fill="both", expand=True, padx=10, pady=10)

        self.scroll_canvas = tk.Canvas(self.timetable_frame)
        self.scroll_frame = tk.Frame(self.scroll_canvas, bg=BACKGROUND_COLOR)
        self.scroll_bar = tk.Scrollbar(self.timetable_frame, orient="vertical", command=self.scroll_canvas.yview)
        self.scroll_canvas.configure(yscrollcommand=self.scroll_bar.set)

        self.scroll_bar.pack(side="right", fill="y")
        self.scroll_canvas.pack(side="left", fill="both", expand=True)
        self.scroll_canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
        self.scroll_frame.bind(
            "<Configure>",
            lambda e: self.scroll_canvas.configure(scrollregion=self.scroll_canvas.bbox("all"))
        )
        
    def get_teachers(self, coordinator_id):

        conn = sqlite3.connect(self.DB_NAME)
        cursor = conn.cursor()
        cursor.execute("SELECT id, name, coordinator_id FROM teachers WHERE coordinator_id = ?", (coordinator_id,))
        
        teachers = cursor.fetchall()
        conn.close()
        return teachers

    def show_modules_screen(self):
        """Expande ou recolhe o painel de módulos."""
        if self.modules_frame.winfo_ismapped():
            self.modules_frame.pack_forget()  
        else:
            self.modules_frame.pack(pady=10, padx=10, fill="x")  

    def show_home_screen(self):
        """Retorna à tela inicial."""
        self.root.destroy()
        home_root = tk.Tk()
        app = ScreenManager(home_root, self.coordinator_id)
        home_root.mainloop()

    def get_coordinator_info(self):
        """Recupera informações do coordenador com base no ID"""
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute("SELECT course FROM coordinators WHERE id = ?", (self.coordinator_id,))
        course = cursor.fetchone()
        conn.close()
        return course[0] if course else None
    
    def get_filtered_classes(self):
        """Retorna apenas as turmas associadas ao coordenador logado."""
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        cursor.execute("SELECT name FROM classes WHERE coordinator_id = ?", (self.coordinator_id,))
        classes = [row[0] for row in cursor.fetchall()]
        
        conn.close()
        return classes

    def get_lab_disciplines_for_class(self, class_name):
        conn = sqlite3.connect(self.DB_NAME)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT d.name
            FROM disciplines d
            JOIN classes c ON d.course = c.name
            WHERE c.name = ? AND d.requires_laboratory = 1 AND d.coordinator_id = ?
        """, (class_name, self.coordinator_id))
        disciplines = [row[0] for row in cursor.fetchall()]
        conn.close()
        return disciplines
    
    def get_lab_name_for_discipline(self, discipline_name):
        conn = sqlite3.connect(self.DB_NAME)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT l.name
            FROM disciplines d
            JOIN laboratories l ON d.laboratory_id = l.id
            WHERE d.name = ? AND d.coordinator_id = ?
        """, (discipline_name, self.coordinator_id))

        result = cursor.fetchone()
        conn.close()

        return result[0] if result else None
    
    def get_discipline_id_by_name(self, discipline_name):
        conn = sqlite3.connect(self.DB_NAME)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT id
            FROM disciplines
            WHERE name = ? AND coordinator_id = ?
        """, (discipline_name, self.coordinator_id))

        result = cursor.fetchone()
        conn.close()

        return result[0] if result else None
    
    def get_all_labs_for_coordinator(self):
        conn = sqlite3.connect(self.DB_NAME)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT name FROM laboratories
            WHERE coordinator_id = ?
            ORDER BY name
        """, (self.coordinator_id,))
        labs = [row[0] for row in cursor.fetchall()]
        conn.close()
        return labs
    
    def get_available_labs_for_class(self, class_name, division_count):
        conn = sqlite3.connect(self.DB_NAME)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT student_count FROM classes
            WHERE name = ? AND coordinator_id = ?
        """, (class_name, self.coordinator_id))
        student_count = cursor.fetchone()

        if not student_count:
            print(f"❌ Classe '{class_name}' não encontrada.")
            conn.close()
            return []

        student_count = student_count[0]

        cursor.execute("""
            SELECT id, name, capacity FROM laboratories
            WHERE coordinator_id = ?
        """, (self.coordinator_id,))
        labs = cursor.fetchall()
        conn.close()

        available_labs = []
        for lab in labs:
            lab_id, lab_name, lab_capacity = lab
            lab_divisions_needed = (student_count // lab_capacity) + (student_count % lab_capacity > 0)

            if lab_capacity >= lab_divisions_needed:
                available_labs.append({'id': lab_id, 'name': lab_name, 'capacity': lab_capacity})

        return available_labs

    def get_lab_capacity(self, lab_id):
        conn = sqlite3.connect(self.DB_NAME)
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT capacity FROM laboratories WHERE id = ? AND coordinator_id = ?
        """, (lab_id, self.coordinator_id))  
        
        result = cursor.fetchone()
        conn.close()
        
        if result:
            return result[0]
        return 0  

    def save_lab_division_config(self, class_name, discipline_name, division_count, selected_lab_name=None):
        conn = sqlite3.connect(self.DB_NAME)
        cursor = conn.cursor()

        class_id = get_class_id(class_name, self.coordinator_id)
        discipline_id = self.get_discipline_id_by_name(discipline_name)

        if class_id is None or discipline_id is None:
            messagebox.showerror("Erro", "Turma ou disciplina não encontrada.")
            conn.close()
            return

        cursor.execute("""
            SELECT id, name, capacity FROM laboratories 
            WHERE coordinator_id = ?
            ORDER BY capacity ASC
        """, (self.coordinator_id,))
        labs = cursor.fetchall()

        cursor.execute("""
            SELECT division_number, lab_id FROM lab_division_config 
            WHERE class_id = ? AND discipline_id = ?
        """, (class_id, discipline_id))
        existing_divisions = cursor.fetchall()
        used_lab_ids = {lab_id for _, lab_id in existing_divisions}
        existing_div_numbers = {div_num for div_num, _ in existing_divisions}

        available_divisions = [i for i in range(1, division_count + 1) if i not in existing_div_numbers]
        if not available_divisions:
            messagebox.showinfo("Info", "Todas as divisões já foram alocadas anteriormente.")
            conn.close()
            return

        available_labs = [lab for lab in labs if lab[0] not in used_lab_ids]
        if not available_labs or len(available_labs) < len(available_divisions):
            messagebox.showerror("Erro", "Não há laboratórios suficientes disponíveis.")
            conn.close()
            return

        allocated_labs = []
        for i, div_number in enumerate(available_divisions):
            lab = available_labs[i % len(available_labs)]
            lab_id, lab_name, _ = lab
            cursor.execute("""
                INSERT INTO lab_division_config (class_id, discipline_id, division_number, lab_id, coordinator_id)
                VALUES (?, ?, ?, ?, ?)
            """, (class_id, discipline_id, div_number, lab_id, self.coordinator_id))
            allocated_labs.append(lab_name)
            used_lab_ids.add(lab_id)

        conn.commit()
        conn.close()

        if allocated_labs:
            unique_labs = sorted(set(allocated_labs))
            messagebox.showinfo("Sucesso", f"Divisões alocadas com sucesso nos laboratórios: {', '.join(unique_labs)}.")

        self.load_lab_config_data()

    def open_lab_config_window(self, event=None):
        window = tk.Toplevel()
        window.title("Configuração de Laboratórios")
        window.geometry("1000x600")
        
        # Main container
        main_frame = ttk.Frame(window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Title
        ttk.Label(main_frame, text="Configuração de Divisões de Laboratório", 
                font=('Helvetica', 12, 'bold')).pack(pady=10)
        
        # Treeview frame
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbars
        y_scroll = ttk.Scrollbar(tree_frame)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        x_scroll = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        x_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Treeview
        columns = ('class', 'students', 'discipline', 'current_divisions', 'configured')
        self.lab_tree = ttk.Treeview(
            tree_frame, 
            columns=columns, 
            show='headings',
            yscrollcommand=y_scroll.set,
            xscrollcommand=x_scroll.set
        )
        
        # Configure columns
        self.lab_tree.heading('class', text='Turma')
        self.lab_tree.heading('students', text='Alunos')
        self.lab_tree.heading('discipline', text='Disciplina')
        self.lab_tree.heading('current_divisions', text='Divisões Atuais')
        self.lab_tree.heading('configured', text='Configurado?')
        
        self.lab_tree.column('class', width=150, anchor=tk.W)
        self.lab_tree.column('students', width=80, anchor=tk.CENTER)
        self.lab_tree.column('discipline', width=250, anchor=tk.W)
        self.lab_tree.column('current_divisions', width=120, anchor=tk.CENTER)
        self.lab_tree.column('configured', width=100, anchor=tk.CENTER)
        
        self.lab_tree.pack(fill=tk.BOTH, expand=True)
        
        y_scroll.config(command=self.lab_tree.yview)
        x_scroll.config(command=self.lab_tree.xview)
        
        # Load initial data
        self.load_lab_config_data()
        
        # Configuration frame
        config_frame = ttk.Frame(main_frame)
        config_frame.pack(fill=tk.X, pady=10)
        
        # Class and discipline selection
        ttk.Label(config_frame, text="Turma:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.class_var = tk.StringVar()
        class_options = self.get_filtered_classes()
        class_dropdown = ttk.Combobox(config_frame, textvariable=self.class_var, values=class_options, state="readonly")
        class_dropdown.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
        
        ttk.Label(config_frame, text="Disciplina:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        self.discipline_var = tk.StringVar()
        self.discipline_dropdown = ttk.Combobox(config_frame, textvariable=self.discipline_var, state="readonly")
        self.discipline_dropdown.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)
        
        ttk.Label(config_frame, text="Divisões:").grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
        self.division_var = tk.StringVar()
        division_entry = ttk.Entry(config_frame, textvariable=self.division_var)
        division_entry.grid(row=2, column=1, padx=5, pady=5, sticky=tk.W)

        class_dropdown.bind("<<ComboboxSelected>>", self.update_lab_disciplines)

        save_btn = ttk.Button(config_frame, text="Salvar Configuração", command=self.save_lab_config)
        save_btn.grid(row=3, column=0, columnspan=2, pady=10)
        
        self.lab_tree.bind("<Double-1>", self.edit_lab_config)

    def save_lab_config(self):
        class_name = self.class_var.get()
        discipline_name = self.discipline_var.get()
        division_count = int(self.division_var.get())

        if not class_name or not discipline_name or division_count <= 0:
            print("❌ Preencha todos os campos corretamente.")
            return

        available_labs = self.get_available_labs_for_class(class_name, division_count)
        if not available_labs:
            print("❌ Não há laboratórios suficientes para alocar as divisões.")
            return

        for lab in available_labs:
            selected_lab_name = lab['name']
            self.save_lab_division_config(class_name, discipline_name, division_count, selected_lab_name)

        print(f"✔️ Divisões alocadas com sucesso nos laboratórios.")

    def load_lab_config_data(self):
        """Load all classes and disciplines that need lab allocation"""
        conn = sqlite3.connect(self.DB_NAME)
        cursor = conn.cursor()
        
        for item in self.lab_tree.get_children():
            self.lab_tree.delete(item)

        cursor.execute("""
            SELECT c.name, c.student_count, d.name, 
                (SELECT COUNT(*) FROM lab_division_config ldc 
                    WHERE ldc.class_id = c.id AND ldc.discipline_id = d.id AND ldc.coordinator_id = ?) as divisions
            FROM classes c
            JOIN disciplines d ON d.course = c.name
            WHERE d.requires_laboratory = 1 AND d.coordinator_id = ?
            ORDER BY c.name, d.name
        """, (self.coordinator_id, self.coordinator_id))
        
        for row in cursor.fetchall():
            class_name, student_count, discipline_name, divisions = row
            configured = "✔" if divisions > 0 else "✖"
            self.lab_tree.insert('', tk.END, 
                                values=(class_name, student_count, discipline_name, divisions, configured),
                                tags=('configured' if configured == "✔" else 'not_configured'))
        
        self.lab_tree.tag_configure('configured', background='#e6ffe6')
        self.lab_tree.tag_configure('not_configured', background='#ffe6e6')       
        conn.close()

    def update_lab_disciplines(self, event=None):
        """Update discipline dropdown based on selected class"""
        selected_class_name = self.class_var.get()
        if not selected_class_name:
            return
        
        disciplines = self.get_lab_disciplines_for_class(selected_class_name)
        self.discipline_dropdown['values'] = disciplines
        self.discipline_var.set("")

    def edit_lab_config(self, event):
        item = self.lab_tree.selection()[0]
        values = self.lab_tree.item(item, 'values')
        
        self.class_var.set(values[0])
        self.discipline_var.set(values[2])
        self.division_var.set(values[3] if values[3] != '0' else "")

    def save_lab_division_config(self, class_name, discipline_name, division_count, selected_lab_name=None):
        conn = sqlite3.connect(self.DB_NAME)
        cursor = conn.cursor()

        # Obter IDs
        class_id = get_class_id(class_name, self.coordinator_id)
        discipline_id = self.get_discipline_id_by_name(discipline_name)

        if class_id is None or discipline_id is None:
            messagebox.showerror("Erro", "Turma ou disciplina não encontrada.")
            conn.close()
            return

        cursor.execute("""
            SELECT id, name, capacity FROM laboratories 
            WHERE coordinator_id = ?
            ORDER BY capacity ASC
        """, (self.coordinator_id,))
        labs = cursor.fetchall()
        if not labs:
            messagebox.showerror("Erro", "Nenhum laboratório disponível.")
            conn.close()
            return

        cursor.execute("""
            SELECT division_number, lab_id FROM lab_division_config 
            WHERE class_id = ? AND discipline_id = ?
        """, (class_id, discipline_id))
        existing_divisions = cursor.fetchall()
        used_lab_ids = {lab_id for _, lab_id in existing_divisions}
        existing_div_numbers = {div_num for div_num, _ in existing_divisions}

        total_needed = division_count
        available_divisions = [i for i in range(1, division_count + 1) if i not in existing_div_numbers]
      
        lab_divisions = []
        divisions_remaining = len(available_divisions)
        cursor.execute("""
            SELECT DISTINCT lab_id 
            FROM lab_division_config 
            WHERE class_id = ? AND coordinator_id = ? AND discipline_id != ?
        """, (class_id, self.coordinator_id, discipline_id))
        labs_used_by_class = {row[0] for row in cursor.fetchall()}

        available_labs = [lab for lab in labs if lab[0] not in used_lab_ids and lab[0] not in labs_used_by_class]

        if not available_labs or len(available_labs) < divisions_remaining:
            messagebox.showerror("Erro", "Não há laboratórios suficientes disponíveis.")
            conn.close()
            return

        allocated_labs = []
        for i, div_number in enumerate(available_divisions):
            lab = available_labs[i % len(available_labs)]
            lab_id, lab_name, _ = lab
            cursor.execute("""
                INSERT INTO lab_division_config (class_id, discipline_id, division_number, lab_id, coordinator_id)
                VALUES (?, ?, ?, ?, ?)
            """, (class_id, discipline_id, div_number, lab_id, self.coordinator_id))
            allocated_labs.append(lab_name)
            used_lab_ids.add(lab_id)

        conn.commit()
        conn.close()

        # Mensagem única
        if allocated_labs:
            unique_labs = sorted(set(allocated_labs))
            messagebox.showinfo("Sucesso", f"Divisões alocadas com sucesso nos laboratórios: {', '.join(unique_labs)}.")
        self.load_lab_config_data()

    def generate_timetable(self):
        conn = sqlite3.connect(self.DB_NAME)
        cursor = conn.cursor()

        disciplines = get_disciplines(self.coordinator_id)
        teachers = get_teacher_data(self.coordinator_id)
        teacher_limits = get_teacher_limits(self.coordinator_id)

        classes = self.get_filtered_classes()
        if not classes:
            print("⚠️ Nenhuma turma encontrada para este coordenador.")
            return {}

        # Estruturas para controle
        timetable = {
            cls: {day: [['', ''] for _ in time_slots] for day in days_of_week}
            for cls in classes
        }
        timetable_entries = {cls: [] for cls in classes}
        teacher_availability = get_teacher_availability_for_timetable(teacher_limits, teachers)

        lab_schedule = {}
        for cls in classes:
            class_id = get_class_id(cls, self.coordinator_id)
            cursor.execute("""
                SELECT l.name FROM lab_division_config ldc
                JOIN laboratories l ON l.id = ldc.lab_id
                WHERE ldc.class_id = ? AND ldc.coordinator_id = ?
            """, (class_id, self.coordinator_id))
            for (lab_name,) in cursor.fetchall():
                if lab_name not in lab_schedule:
                    lab_schedule[lab_name] = {day: set() for day in days_of_week}

        
        # Dicionários para rastrear alocações
        teacher_load = {t: 0 for t in teachers}  
        teacher_schedule = {t: {day: [] for day in days_of_week} for t in teachers} 

        for cls in classes:
            class_course = get_class_course(cls, self.coordinator_id)
            class_disciplines = [d for d in disciplines if d['course'] == class_course]

            discipline_hours = {d['name']: d['hours'] for d in class_disciplines}
            assigned_hours = {d['id']: 0 for d in class_disciplines}

            class_id = get_class_id(cls, self.coordinator_id)

            cursor.execute("""
                SELECT ldc.discipline_id, ldc.division_number, l.name 
                FROM lab_division_config ldc
                JOIN laboratories l ON ldc.lab_id = l.id
                WHERE ldc.class_id = ? AND ldc.coordinator_id = ?
            """, (class_id, self.coordinator_id))

            lab_divisions = cursor.fetchall()
            lab_division_map = {}
            for discipline_id, division_number, lab_name in lab_divisions:
                lab_division_map.setdefault(discipline_id, []).append({
                    "division_number": division_number,
                    "lab_name": lab_name
                })

            for day in days_of_week:
                for i, time_slot in enumerate(time_slots):
                    if time_slot == "20:25 - 20:45":
                        continue  # intervalo

                    inicio, termino = time_slot.split(" - ")
                    
                    # Disciplinas que ainda precisam de horas
                    possible_disciplines = [
                        d for d in class_disciplines 
                        if assigned_hours[d['id']] < discipline_hours[d['name']]
                    ]
                    
                    if not possible_disciplines:
                        continue

                    discipline_obj = random.choice(possible_disciplines)
                    discipline_name = discipline_obj['name']
                    discipline_id = discipline_obj['id']
                    requires_lab = discipline_obj.get('requires_laboratory', False)

                    # Encontrar professor disponível
                    available_teachers = [
                        t for t in teachers 
                        if day in teacher_availability.get(t, []) and
                        time_slot not in teacher_schedule[t][day]
                    ]

                    # Ordenar por professores com menos aulas alocadas
                    available_teachers.sort(key=lambda t: teacher_load[t])
                    teacher = available_teachers[0] if available_teachers else None

                    if time_slot in lab_schedule[lab_name][day]:
                        continue  # horário já ocupado neste lab

                    if requires_lab and discipline_id in lab_division_map:
                        # Alocar todas as divisões obrigatórias de laboratório
                        for division in lab_division_map[discipline_id]:
                            lab_name = division["lab_name"]
                            division_number = division["division_number"]
                            disciplina_label = f"{discipline_name} - D{division_number}"

                            alocado = False
                            for day in days_of_week:
                                for i, time_slot in enumerate(time_slots):
                                    if time_slot == "20:25 - 20:45":
                                        continue  # intervalo
                                    if timetable[cls][day][i][0]:  # já ocupado
                                        continue
                                    # Verificar professor disponível
                                    available_teachers = [
                                        t for t in teachers 
                                        if day in teacher_availability.get(t, []) and
                                        time_slot not in teacher_schedule[t][day]
                                    ]
                                    available_teachers.sort(key=lambda t: teacher_load[t])
                                    teacher = available_teachers[0] if available_teachers else None

                                    # Registrar aula de laboratório
                                    entry = {
                                        "DIA": day,
                                        "INÍCIO": time_slot.split(" - ")[0],
                                        "TÉRMINO": time_slot.split(" - ")[1],
                                        "CÓDIGO": f"CMP{i+1:03}-D{division_number}",
                                        "NOME": discipline_name,
                                        "TURMA LAB": lab_name,
                                        "PROFESSOR": teacher or "A definir",
                                        "TEÓRICA": "",
                                        "PRÁTICA": "X",
                                        "ENCONTRO": ""
                                    }

                                    timetable[cls][day][i] = [disciplina_label, teacher or "A definir"]
                                    timetable_entries[cls].append(entry)
                                    assigned_hours[discipline_id] += 1
                                    if teacher:
                                        teacher_load[teacher] += 1
                                        teacher_schedule[teacher][day].append(time_slot)
                                    alocado = True
                                    break
                                if alocado:
                                    lab_schedule[lab_name][day].add(time_slot)
                                    break

                            if not alocado:
                                print(f"⚠️ Aula de laboratório '{disciplina_label}' não foi alocada!")

                    # Processar aula teórica
                    disciplina_label = discipline_name
                    entry = {
                        "DIA": day,
                        "INÍCIO": inicio,
                        "TÉRMINO": termino,
                        "CÓDIGO": f"CMP{i+1:03}",
                        "NOME": discipline_name,
                        "TURMA LAB": "",
                        "PROFESSOR": teacher or "A definir",
                        "TEÓRICA": "X",
                        "PRÁTICA": "",
                        "ENCONTRO": ""
                    }

                    assigned_hours[discipline_id] += 1
                    if teacher:
                        teacher_load[teacher] += 1
                        teacher_schedule[teacher][day].append(time_slot)
                    timetable[cls][day][i] = [disciplina_label, teacher or "A definir"]
                    timetable_entries[cls].append(entry)

        conn.close()
        
        # Verificação final
        undefined_teachers = sum(1 for cls in timetable_entries for entry in timetable_entries[cls] if entry["PROFESSOR"] == "A definir")
        total_classes = sum(len(entries) for entries in timetable_entries.values())
        
        if undefined_teachers > 0:
            print(f"⚠️ Aviso: {undefined_teachers} de {total_classes} aulas sem professor alocado")
            print("Possíveis soluções:")
            print("1. Verifique a disponibilidade dos professores")
            print("2. Contrate mais professores")
            print("3. Redistribua as disciplinas entre os professores existentes")
        
        return timetable_entries

    def show_timetable(self):
        for widget in self.scroll_frame.winfo_children():
            widget.destroy()
            
        self.timetable = self.generate_timetable()
            
        for name, timetable_class in self.timetable.items():
            self.create_class_table(self.scroll_frame, name, timetable_class)

    def create_class_table(self, frame, class_name, timetable_class):
        week_order = {
            "Segunda": 0,
            "Terça": 1,
            "Quarta": 2,
            "Quinta": 3,
            "Sexta": 4,
        }

        timetable_class.sort(key=lambda x: (week_order.get(x["DIA"], 99), x["INÍCIO"]))

        class_frame = tk.Frame(frame)
        class_frame.pack(pady=20, fill="x", expand=True)

        title = tk.Label(
            class_frame,
            text=f"TURMA {class_name}",
            font=("Helvetica", 12, "bold"),
            pady=10
        )
        title.grid(row=0, column=0, columnspan=11, sticky="w")

        var = tk.BooleanVar()
        checkbox = tk.Checkbutton(
            class_frame,
            text="",
            variable=var,
            command=lambda: self.select_grade(class_name, var)
        )
        checkbox.grid(row=0, column=11, padx=20, sticky="e")

        headers = ["DIA", "INÍCIO", "TÉRMINO", "CÓDIGO", "NOME", "TURMA LAB", "PROFESSOR", "TEÓRICA", "PRÁTICA", "ENCONTRO"]

        for col in range(len(headers)):
            class_frame.grid_columnconfigure(col, weight=1) 

        for col, header in enumerate(headers):
            label = tk.Label(
                class_frame,
                text=header,
                font=("Helvetica", 10, "bold"),
                relief="ridge",
                borderwidth=1,
                bg="#F5F5F5",
                padx=5,
                pady=3,
                anchor="center"
            )
            label.grid(row=1, column=col, sticky="nsew")

        row = 2
        i = 0
        while i < len(timetable_class):
            current = timetable_class[i]
            day = current["DIA"]

            rowspan = 1
            for j in range(i + 1, len(timetable_class)):
                if timetable_class[j]["DIA"] == day:
                    rowspan += 1
                else:
                    break

            day_label = tk.Label(
                class_frame,
                text=day,
                font=("Helvetica", 9),
                relief="ridge",
                borderwidth=1,
                padx=5,
                pady=2,
                bg="#E0E0E0",
                anchor="nw",
                justify="left"
            )
            day_label.grid(row=row, column=0, rowspan=rowspan, sticky="nsew")

            for k in range(rowspan):
                entry = timetable_class[i + k]
                bg_color = "#FFFFFF" if (row + k) % 2 == 0 else "#F0F0F0"

                for col_idx, header in enumerate(headers[1:], start=1):
                    value = entry.get(header, "")
                    label = tk.Label(
                        class_frame,
                        text=value,
                        font=("Helvetica", 9),
                        relief="ridge",
                        borderwidth=1,
                        padx=5,
                        pady=2,
                        anchor="nw",
                        justify="left",
                        bg=bg_color,
                        wraplength=150 
                    )
                    label.grid(row=row + k, column=col_idx, sticky="nsew")

            i += rowspan
            row += rowspan

    def select_grade(self, grade_name, var):
        if var.get():
            if grade_name not in self.selected_grades:
                self.selected_grades.append(grade_name)
        else:
            if grade_name in self.selected_grades:
                self.selected_grades.remove(grade_name)

        if self.selected_grades:
            self.download_button.config(state="normal")
        else:
            self.download_button.config(state="disabled")
            

    def show_export_options(self):
        if not self.selected_grades:
            messagebox.showwarning("Aviso", "Nenhuma grade selecionada para download.")
            return
        
        option_window = tk.Toplevel(self.root)
        option_window.title("Escolher formato de exportação")
        option_window.geometry("300x150")
        option_window.grab_set() 

        excel_var = tk.BooleanVar(value=True)
        pdf_var = tk.BooleanVar(value=False)

        tk.Checkbutton(option_window, text="Exportar para Excel (.xlsx)", variable=excel_var).pack(pady=5)
        tk.Checkbutton(option_window, text="Exportar para PDF (.pdf)", variable=pdf_var).pack(pady=5)

        def start_export():
            export_excel = excel_var.get()
            export_pdf = pdf_var.get()
            option_window.destroy()
            self.download_grade(export_excel, export_pdf)

        tk.Button(option_window, text="Exportar", command=start_export).pack(pady=10)

    DB_NAME = "schedule.db"

    def download_grade(self, export_excel=True, export_pdf=False):
        if not self.selected_grades:
            messagebox.showwarning("Aviso", "Nenhuma grade selecionada para download.")
            return

        for grade_name in self.selected_grades:
            dir_path = filedialog.askdirectory(title="Escolha uma pasta para salvar os arquivos")
            if not dir_path:
                continue

            entries = self.manual_timetable.get(grade_name) or self.timetable.get(grade_name, [])
            headers = ["DIA", "INÍCIO", "TÉRMINO", "CÓDIGO", "NOME", "TURMA LAB", "PROFESSOR", "TEÓRICA", "PRÁTICA", "ENCONTRO"]

            grade_content = f"Grade de {grade_name}\n"
            grade_content += "\t".join(headers) + "\n"
            for entry in entries:
                row = "\t".join([str(entry.get(h, "")) for h in headers])
                grade_content += row + "\n"
            grade_content += "\n" + "=" * 30 + "\n"

            # Exporta para Excel
            if export_excel:
                excel_path = os.path.join(dir_path, f"{grade_name}.xlsx")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = grade_name[:31]
                ws.append(headers)
                for entry in entries:
                    ws.append([entry.get(h, "") for h in headers])
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    max_length = max(len(str(cell.value)) for cell in ws[get_column_letter(col)])
                    ws.column_dimensions[get_column_letter(col)].width = max_length + 2
                wb.save(excel_path)
                print(f"✅ Excel salvo: {excel_path}")

            # Exporta para PDF
            if export_pdf:
                pdf_path = os.path.join(dir_path, f"{grade_name}.pdf")
                c = canvas.Canvas(pdf_path, pagesize=A4)
                width, height = A4
                y = height - 50
                c.setFont("Helvetica-Bold", 14)
                c.drawString(50, y, f"Grade de {grade_name}")
                y -= 30
                c.setFont("Helvetica", 10)
                for entry in entries:
                    line = " | ".join([f"{h}: {entry.get(h, '')}" for h in headers])
                    c.drawString(50, y, line[:180])
                    y -= 15
                    if y < 50:
                        c.showPage()
                        y = height - 50
                c.save()
                print(f"📄 PDF salvo: {pdf_path}")

            # Salva no banco
            conn = sqlite3.connect(self.DB_NAME)
            cursor = conn.cursor()
            cursor.execute("""
            INSERT INTO saved_grades (name, content, file_path, coordinator_id)
            VALUES (?, ?, ?, ?)
            """, (grade_name, grade_content, dir_path, self.coordinator_id))
            conn.commit()
            conn.close()

        messagebox.showinfo("Sucesso", "Exportação concluída e grades salvas no banco!")

    def export_timetable_to_excel(self, timetable, filename="horario_gerado.xlsx"):
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            for cls, entries in timetable.items():
                ws = wb.create_sheet(title=cls[:31])

                headers = ["DIA", "INÍCIO", "TÉRMINO", "CÓDIGO", "NOME", "TURMA LAB", "PROFESSOR", "TEÓRICA", "PRÁTICA", "ENCONTRO"]
                ws.append(headers)

                for entry in entries:
                    ws.append([
                        entry.get("DIA", ""),
                        entry.get("INÍCIO", ""),
                        entry.get("TÉRMINO", ""),
                        entry.get("CÓDIGO", ""),
                        entry.get("NOME", ""),
                        entry.get("TURMA LAB", ""),
                        entry.get("PROFESSOR", ""),
                        entry.get("TEÓRICA", ""),
                        entry.get("PRÁTICA", ""),
                        entry.get("ENCONTRO", "")
                    ])

                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    max_length = max(len(str(c.value)) if c.value else 0 for c in ws[get_column_letter(col)])
                    ws.column_dimensions[get_column_letter(col)].width = max_length + 2

            wb.save(filename)
            print(f"✅ Arquivo exportado: {filename}")


    def select_cell(self, day, time_slot, label):
            if self.selected_cell:
                self.selected_cell.config(bg=WHITE_COLOR)
            
            self.selected_cell = label
            self.selected_cell.config(bg=LABEL_COLOR)
            
            text = self.selected_cell.cget("text")
            self.original_discipline = text.split("\n")[0]  
            self.original_teacher = text.split("\n")[1]   
            
            self.edit_button.config(state="normal")
            self.save_button.config(state="normal")
            self.cancel_button.config(state="normal")
            self.delete_button.config(state="normal")

    def edit_cell(self):
        if not self.selected_cell:
            return
        
        self.original_teacher = self.selected_cell.cget("text").split("\n")[1]
        self.original_discipline = self.selected_cell.cget("text").split("\n")[0]
        self.open_edit_modal()

    def open_edit_modal(self):
        modal_window = tk.Toplevel(self.root)
        modal_window.title("Editar Disciplina e Professor")

        label_discipline = tk.Label(modal_window, text="Selecione uma nova disciplina:")
        label_discipline.pack(pady=4)

        discipline_select = ttk.Combobox(modal_window, values=[d['name'] for d in get_disciplines(self.coordinator_id)])
        discipline_select.set(self.original_discipline)
        discipline_select.pack(pady=4)

        label_teacher = tk.Label(modal_window, text="Selecione um novo professor:")
        label_teacher.pack(pady=4)

        teacher_select = ttk.Combobox(modal_window, values=list(teachers.keys()))
        teacher_select.set(self.original_teacher)
        teacher_select.pack(pady=4)

        def save_changes():
            new_discipline = discipline_select.get()
            new_teacher = teacher_select.get()
            self.selected_cell.config(text=f"{new_discipline}\n{new_teacher}")

            for cls, schedule in self.timetable.items():
                for day, slots in schedule.items():
                    for i, (discipline, teacher) in enumerate(slots):
                        if (discipline, teacher) == (self.original_discipline, self.original_teacher):
                            self.timetable[cls][day][i] = [new_discipline, new_teacher]

            modal_window.destroy()
            self.save_button.config(state="disabled")
            self.cancel_button.config(state="disabled")

        def cancel():
            modal_window.destroy()

        save_button = tk.Button(modal_window, text="Salvar", command=save_changes)
        save_button.pack(pady=4)
        
        cancel_button = tk.Button(modal_window, text="Cancelar", command=cancel)
        cancel_button.pack(pady=4)

    def save_changes(self):
        if self.selected_cell:
            self.save_button.config(state="disabled")
            self.cancel_button.config(state="disabled")

    def cancel_edit(self):
        if self.selected_cell:
            self.selected_cell.config(text=f"{self.original_discipline}\n{self.original_teacher}")
        
        self.save_button.config(state="disabled")
        self.cancel_button.config(state="disabled")
        
    def confirm_delete_schedule(self):
        if self.selected_cell:
            confirmation = askyesno("Confirmar Exclusão", "Você tem certeza que deseja excluir este registro?")
            if confirmation:
                self.delete_schedule()
    
    def delete_schedule(self):
        if self.selected_cell:
            self.selected_cell.config(text="[REMOVIDO]")
            self.selected_cell.config(bg=WHITE_COLOR)
            self.selected_cell = None
            self.edit_button.config(state="disabled")
            self.delete_button.config(state="disabled")
            self.save_button.config(state="disabled")
            self.cancel_button.config(state="disabled")

    def create_manual_schedule(self):
        manual_schedule_window = tk.Toplevel(self.root)
        manual_schedule_window.title("Criar Grade Manualmente")
        manual_schedule_window.geometry("800x600")

        class_label = tk.Label(manual_schedule_window, text="Escolha a Turma:")
        class_label.pack(pady=8)

        classes = self.get_filtered_classes()
        class_var = tk.StringVar()
        class_select = ttk.Combobox(manual_schedule_window, textvariable=class_var, values=classes, state="readonly")
        class_select.pack(pady=8)

        table_frame = tk.Frame(manual_schedule_window)
        table_frame.pack(pady=10, fill="both", expand=True)

        discipline_data = get_disciplines(self.coordinator_id)
        teacher_data = get_teacher_data(self.coordinator_id)
        teachers = list(set(teacher_data))
        entries = []

        def load_schedule_table(event=None):
            for widget in table_frame.winfo_children():
                widget.destroy()
            selected_class = class_var.get()
            class_course = get_class_course(selected_class, self.coordinator_id)
            class_disciplines = [d for d in discipline_data if d['course'] == class_course]

            tk.Label(table_frame, text="DIA").grid(row=0, column=0)
            tk.Label(table_frame, text="HORÁRIO").grid(row=0, column=1)
            tk.Label(table_frame, text="DISCIPLINA").grid(row=0, column=2)
            tk.Label(table_frame, text="PROFESSOR").grid(row=0, column=3)

            row_idx = 1
            entries.clear()

            for day in days_of_week:
                for slot in time_slots:
                    if slot == "20:25 - 20:45":
                        continue

                    tk.Label(table_frame, text=day).grid(row=row_idx, column=0)
                    tk.Label(table_frame, text=slot).grid(row=row_idx, column=1)

                    discipline_var = tk.StringVar()
                    discipline_cb = ttk.Combobox(table_frame, textvariable=discipline_var, values=[d["name"] for d in class_disciplines])
                    discipline_cb.grid(row=row_idx, column=2)

                    teacher_var = tk.StringVar()
                    teacher_cb = ttk.Combobox(table_frame, textvariable=teacher_var, values=teachers)
                    teacher_cb.grid(row=row_idx, column=3)

                    entries.append({
                        "day": day,
                        "slot": slot,
                        "discipline_var": discipline_var,
                        "teacher_var": teacher_var
                    })

                    row_idx += 1

        class_select.bind("<<ComboboxSelected>>", load_schedule_table)

        def save_manual_schedule():
            selected_class = class_var.get()
            if not selected_class:
                messagebox.showwarning("Aviso", "Selecione uma turma.")
                return

            schedule = []
            used_slots = set()
            teacher_usage = {}

            for entry in entries:
                day = entry["day"]
                slot = entry["slot"]
                discipline = entry["discipline_var"].get()
                teacher = entry["teacher_var"].get()

                if not discipline or not teacher:
                    continue

                key = (day, slot)
                if key in used_slots:
                    messagebox.showerror("Erro", f"Conflito de horário detectado: {day}, {slot}")
                    return

                if (teacher, day, slot) in teacher_usage:
                    messagebox.showerror("Erro", f"Professor {teacher} já está alocado em {day} às {slot}.")
                    return

                used_slots.add(key)
                teacher_usage[(teacher, day, slot)] = True

                discipline_obj = next((d for d in discipline_data if d["name"] == discipline), None)
                pratica = discipline_obj.get("requires_laboratory", False) if discipline_obj else False

                schedule.append({
                    "DIA": day,
                    "INÍCIO": slot.split(" - ")[0],
                    "TÉRMINO": slot.split(" - ")[1],
                    "CÓDIGO": "",
                    "NOME": discipline,
                    "TURMA LAB": selected_class if pratica else "",
                    "PROFESSOR": teacher,
                    "TEÓRICA": "" if pratica else "X",
                    "PRÁTICA": "X" if pratica else "",
                    "ENCONTRO": ""
                })

            # 🚨 VERIFICAÇÕES DE DISCIPLINAS PRÁTICAS E CONFIGURAÇÃO DE LABORATÓRIO
            for d in schedule:
                if d["PRÁTICA"] == "X":
                    disciplina_nome = d["NOME"]
                    disciplina_obj = next((disc for disc in discipline_data if disc["name"] == disciplina_nome), None)
                    if disciplina_obj:
                        disciplina_id = disciplina_obj["id"]
                        divisoes = self.lab_division_map.get(disciplina_id, 0)
                        if divisoes == 0:
                            messagebox.showerror("Erro", f"Disciplina '{disciplina_nome}' requer laboratório, mas não possui divisões configuradas.")
                            return

            grade_name = f"{selected_class}_MANUAL"
            self.manual_timetable[grade_name] = schedule
            self.selected_grades = [grade_name]

            manual_schedule_window.destroy()
            self.show_export_options()

        save_button = tk.Button(manual_schedule_window, text="Salvar Grade", command=save_manual_schedule, bg="#A0D6B4")
        save_button.pack(pady=10)

        lab_config_button = tk.Button(manual_schedule_window, text="Configurar Laboratórios", command=self.open_lab_config_window, bg="#FFD580")
        lab_config_button.pack(pady=5)

if __name__ == "__main__":
    root = tk.Tk()
    app = TimetableApp(root)
    root.mainloop()
